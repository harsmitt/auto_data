from .tranform_data import *
import openpyxl
from django.http import HttpResponse
# from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from .inline_dump_functions import *
from openpyxl.styles import Color, Fill,Font,Alignment
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side


class excl_format():
    alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=1)
    alignment2 = Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=2)
    alignment3 = Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=3)

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    my_green = openpyxl.styles.colors.Color(rgb='70AB47')
    my_g1 = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_green)
    my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
    my_r1 = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
    my_yel = openpyxl.styles.colors.Color(rgb='00FFF2CC')
    my_y1 = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_yel)
    yellow_1 = openpyxl.styles.colors.Color(rgb="00FFFF00")
    my_y2 = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=yellow_1)


format_excel =excl_format




def set_allign(**kwargs):
    if 'value' in kwargs:
        print(kwargs['value'])
        kwargs['sheet']['A' + str(kwargs['row_no'])].value = kwargs['value']

    if 'allign' in kwargs:
        kwargs['sheet']['A' + str(kwargs['row_no'])].alignment = kwargs['allign']
    if 'color' in kwargs:
        kwargs['sheet']['A' + str(kwargs['row_no'])].fill = kwargs['color']


    kwargs['sheet']['A' + str(kwargs['row_no'])].font = kwargs['sheet']['A' + str(kwargs['row_no'])].font.copy(
        name='Arial')
    kwargs['sheet']['A' + str(kwargs['row_no'])].border = format_excel.thin_border

    return kwargs['sheet']


def subsec_write(**kwargs):
    for s_key, p_key in kwargs['data'].items():
        kwargs['s_start_col'].append(kwargs['row_no'])
        start_col = ''
        end_col = ''
        if type(kwargs['data'][s_key]) == OrderedDict:
            f_row = kwargs['row_no']
            # add subsection key
            kwargs['sheet']= set_allign(sheet =kwargs['sheet'],row_no=kwargs['row_no'],value=s_key,allign =format_excel.alignment)
            ch = 66
            kwargs['row_no'] += 1
            start_col = kwargs['row_no']

            kwargs['sheet'],kwargs['row_no'],end_col = pdf_key_write(data=p_key,sheet=kwargs['sheet'],\
                                                                     row_no=kwargs['row_no'],\
                                                                     end_col=end_col,loop_key=kwargs['loop_key'])

            ch = 66
            for d1 in kwargs['date_list']:
                if start_col and end_col and start_col != end_col:
                    kwargs['sheet'][chr(ch) + str(f_row)].value = '=SUM(' + (chr(ch) + str(start_col)) + ':' + (
                        chr(ch) + str(end_col)) + ')'
                else:
                    kwargs['sheet'][chr(ch) + str(f_row)].value = 0
                kwargs['sheet'][chr(ch) + str(f_row)].fill = format_excel.my_y1
                kwargs['sheet'][chr(ch) + str(f_row)].border = format_excel.thin_border

                # kwargs['sheet'] = set_allign(sheet=kwargs['sheet'], row_no=kwargs['row_no'],color=format_excel.my_y1)

                ch += 1

        else:
            kwargs['sheet'] = set_allign(sheet=kwargs['sheet'], row_no=kwargs['row_no'], value=s_key,
                                         allign=format_excel.alignment)
            ch = 66
            s2_col = []

            s2_total=kwargs['row_no']
            kwargs['row_no'] += 1
            for s2_sec in kwargs['data'][s_key]:
                for s2_k, s2_val in s2_sec.items():
                    s2_col.append(kwargs['row_no'])
                    s2start_col = ''
                    s2end_col = ''
                    s2f_row = kwargs['row_no']
                    kwargs['sheet'] = set_allign(sheet=kwargs['sheet'], row_no=kwargs['row_no'], value=s2_k,
                                                 allign=format_excel.alignment2)
                    ch = 66

                    kwargs['row_no'] += 1
                    s2start_col = kwargs['row_no']
                    kwargs['sheet'], kwargs['row_no'], s2end_col = pdf_key_write(data=s2_val, \
                                                                                 sheet=kwargs['sheet'], \
                                                                               row_no=kwargs['row_no'], \
                                                                               end_col=s2end_col,
                                                                               loop_key=kwargs['loop_key'])

                    ch = 66
                    for d1 in kwargs['date_list']:
                        if s2start_col and s2end_col:
                            if s2start_col != s2end_col:
                                kwargs['sheet'][chr(ch) + str(s2f_row)].value = '=SUM(' + (chr(ch) + str(s2start_col)) + ':' + (
                                    chr(ch) + str(s2end_col)) + ')'
                            else:
                                kwargs['sheet'][chr(ch) + str(s2f_row)].value = '=SUM(' + (chr(ch) + str(s2start_col)) + ')'
                        else:
                            kwargs['sheet'][chr(ch) + str(s2f_row)].value = 0

                        kwargs['sheet'][chr(ch) + str(s2f_row)].fill = format_excel.my_y1
                        kwargs['sheet'][chr(ch) + str(s2f_row)].border = format_excel.thin_border

                        ch += 1
                ch = 66
                for d1 in kwargs['date_list']:
                    str1 = ''
                    for r1 in s2_col:
                        str1 += (chr(ch) + str(r1)) + ','
                    kwargs['sheet'][chr(ch) + str(s2_total)].value = '=SUM(' + str1 + ')'
                    kwargs['sheet'][chr(ch) + str(s2_total)].fill = format_excel.my_g1
                    kwargs['sheet'][chr(ch) + str(s2_total)].border = format_excel.thin_border
                    ch += 1
    return kwargs['sheet'],kwargs['row_no'],kwargs['s_start_col']






def pdf_key_write(**kwargs):
    for p_key1, p_val in kwargs['data'].items():
        # add pdf_key
        kwargs['sheet'] = set_allign(sheet=kwargs['sheet'], row_no=kwargs['row_no'], value=p_key1,
                                     allign=format_excel.alignment2, color=format_excel.my_r1)

        for p_val1, p_val2 in p_val.items():
            ch = 66
            for d1, d2 in kwargs['loop_key'].items():
                if d1 == p_val1:
                    kwargs['sheet'][chr(ch) + str(kwargs['row_no'])].value = p_val2
                # else:
                #     kwargs['sheet'][chr(ch) + str(kwargs['row_no'])].value = 0
                kwargs['sheet'][chr(ch) + str(kwargs['row_no'])].fill = format_excel.my_r1
                kwargs['sheet'][chr(ch) + str(kwargs['row_no'])].border = format_excel.thin_border

                ch += 1
        kwargs['end_col'] = kwargs['row_no']
        kwargs['row_no'] += 1

    return kwargs['sheet'],kwargs['row_no'],kwargs['end_col']