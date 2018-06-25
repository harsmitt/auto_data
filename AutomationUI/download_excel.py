# import pandas
# from openpyxl import load_workbook
#
# book = load_workbook('/home/administrator/16030102.xlsx')
# writer = pandas.ExcelWriter('/home/administrator/16030102.xlsx', engine='openpyxl')
# writer.book = book
# writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#
# # df.to_excel(writer, "Main", cols=['Diff1', 'Diff2'])
# df.to_excel(writer, sheet_name='sheet1', index=False)
#
# writer.save()

from BalanceSheet.models import *
from PNL.models import *
from DataExtraction.common_files.utils import *
import openpyxl
from django.http import HttpResponse
# from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook


date_obj = qtr_date_pnl()
date_obj.update(year_date('December'))
q_list = list(qtr_date_pnl().keys())[-5:]
y_list = list(year_date('December').keys())[-4:]

fill_qy=q_list+y_list
row_mapping = {'Current Assets':65, 'Non-Current Assets':79, 'Current Liabilities':96, 'Non-Current Liabilities':108, 'Shareholder Equity':121}
other_mapping ={'Other Current Assets':70, 'Other Non-Current Assets':86, 'Other Current Liabilities':100, 'Other Non-Current Liabilities':111}

pnl_row_mapping = {'Revenue':14, 'Cost of Revenue': 19,'Other Operating Income': 25,
                   'Operating Expenses':29,'Non-Operating Income/(Expenses)':36,
                   'Income Tax Expense':45, 'Profit/(Loss) from Discontinued Operations':49,
                   'Net Profit/(Loss) for the Year':53,'Depreciation & Dividend':57}

def download_pdf(request):
    wb = openpyxl.load_workbook('/home/administrator/DITMT-13010101.6-Air Conditioning Freezing  Heating Equipment Manufacturing..QC2.xlsx')
    sheet = wb.get_sheet_by_name('BasicInfo')
    for i in range(1,18):
        sheet['C'+str(i)].value=''
        sheet['D' + str(i)].value = ''
    sheet['C18'].value = date_obj[q_list[-1]].title()
    sheet['C19'].value = date_obj[q_list[-1]].title()
    sheet['D18'].value =''
    sheet['D19'].value = ''
    for i in range(20,27):
        sheet['C' + str(i)].value = ''
        sheet['D' + str(i)].value = ''
        # sheet['D18'].value = qtr_list['lrq']
        # sheet['C18'].value = year_list['y4']
    from django.db.models import Q
    sheet1 = wb.get_sheet_by_name('User_Financial_Input')
    b_data = quarter_data.objects.filter(company_name_id=int(request.GET['c_id']),page_extraction='bsheet')
    sec_list = Section.objects.filter(i_related='Balance Sheet')

    pnl_data = quarter_data.objects.filter(company_name_id=int(request.GET['c_id']),page_extraction='pnl')
    pnl_sec_list = Section.objects.filter(i_related='Profit and Loss')

    for sec in sec_list:
        if sec.item != "Shareholder Equity":
            b_part1_data = b_data.filter(Q(section= sec),~Q(subsection=None),~Q(subsection__item__icontains='Other'),~Q(subsection__item__icontains='Deduction'))
            b_part1_s2_data = b_data.filter(Q(section= sec),Q(subsection__item__contains='Other'), ~Q(s2section=None))
            row = fill_data(row_mapping[sec.item],b_part1_data,sheet1)
            row = fill_data(other_mapping['Other '+sec.item], b_part1_s2_data, sheet1,sec_type ='s2')
        else:
            b_part1_data = b_data.filter(Q(section=sec), ~Q(subsection=None), ~Q(subsection__item__icontains='Deduction'))
            row = fill_data(row_mapping[sec.item], b_part1_data, sheet1)
    # #
    for sec in pnl_sec_list:
        if sec.item == 'Extra PNL Keywords':
            pass
        else:
            p_data = pnl_data.filter(Q(section=sec), ~Q(subsection=None))
            if p_data:
                row = fill_data(pnl_row_mapping[sec.item], p_data, sheet1)

    # wb.save('/home/administrator/DataAutomation/company_pdf/different patterns/aditi/DITMT-13010101.6-Air Conditioning Freezing  Heating Equipment Manufacturing..QC2.xlsx')
    c_name= CompanyList.objects.get(id= int(request.GET['c_id']))
    print (c_name)#.values_list('ditname__dit_name',flat=True)
    # response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    # response['Content-Disposition'] = 'attachment; filename='+c_name[0]+'.xlsx'
    stream = save_virtual_workbook(wb)
    response = HttpResponse(stream, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=%s.xlsx' %(c_name.company_name)
    # wb.save(response)
    return response

def get_val(obj):
    val=0
    if obj.description:
        for i in obj.description.split('##'):
            val += int(get_digit(i, ui_num=True))
    return val

def fill_data(row,data_objs,sheet,sec_type=None):
    if sec_type=='s2':
        sub_name = list(data_objs.values_list('s2section__item', flat=True).distinct())
    else:
        sub_name = list(data_objs.values_list('subsection__item', flat=True).distinct())
    for sub in sub_name:
        if sec_type=='s2':
            sub_objs = data_objs.filter(s2section__item=sub)
        else:
            sub_objs = data_objs.filter(subsection__item = sub)
        ch = 70
        for obj in sub_objs:
            print (obj.quarter_date)
            for da in fill_qy:
                if obj.quarter_date == str(date_obj[da]):
                    obj_val = get_val(obj)
                    sheet[chr(ch)+str(row)].value =obj_val
                    print (chr(ch) ,obj.quarter_date )
                    ch += 1
                    break;

        row+=1

        print(row)
    return sheet
        #
